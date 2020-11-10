from openpyxl import load_workbook #Excel情報の読書
import datetime #日付情報の取得

#Excelファイルのパス(.xlsxかxlsmのみ)
filepath = 'C:\\Users\\artmi\\OneDrive\\ドキュメント\\python\\app\\Book1.xlsx'
#Excelの編集の準備
wb = load_workbook(filename=filepath)
#Excel内のシート名
ws1 = wb['データ']
ws2 = wb['集計']

#データの開始位置と終了位置
startdate = datetime.datetime(int(ws2['B2'].value), int(ws2['C2'].value), int(ws2['D2'].value))
enddate = datetime.datetime(int(ws2['B3'].value), int(ws2['C3'].value), int(ws2['D3'].value))

#Excelの最終行と最終列を取得
lastrow1 = ws1.max_row
lastrow2 = ws2.max_row
lastcol2 = ws2.max_column

#シートから情報を読み取る
values1 = [[cell.value for cell in row1] for row1 in ws1]

for i in range(7, lastrow2+1):
    for j in range(2, lastcol2+1):
        counter = 0
        for k in range(1, lastrow1):
            if  values1[k][1] == ws2.cell(row=i, column=1).value:
                if values1[k][2] == ws2.cell(row=6, column=j).value:
                    torihikidate = values1[k][3]
                    if startdate <= torihikidate <= enddate:
                        kingaku=values1[k][4]
                        counter = counter + int(kingaku)
        if counter is None:
            counter = 0
        ws2.cell(row=i, column=j).value = counter

newfilepath = 'C:\\Users\\artmi\\OneDrive\\ドキュメント\\python\\app\\Book2.xlsx'
wb.save(newfilepath)